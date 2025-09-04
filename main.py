# main.py
from __future__ import annotations

from datetime import date
from typing import Optional
import io
import re
import csv
from io import BytesIO, StringIO

import pandas as pd
from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, File, UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

import models, schemas
from database import engine, get_db
from contextlib import asynccontextmanager

# Auth (fastapi-users)
from auth import (
    fastapi_users,
    auth_backend,
    get_current_active_user,
    get_current_active_admin,
    UserRead, UserCreate, UserUpdate,
)

# ==============================
# FastAPI app & lifecycle
# ==============================
@asynccontextmanager
async def lifespan(app: FastAPI):
    # startup: crea todas las tablas
    models.Base.metadata.create_all(bind=engine)
    yield
    # shutdown: nada por ahora

app = FastAPI(title="Dashboard J&J", lifespan=lifespan)

# Estáticos
app.mount("/static", StaticFiles(directory="static"), name="static")

# Routers de auth
app.include_router(
    fastapi_users.get_auth_router(auth_backend),
    prefix="/auth/jwt",
    tags=["auth"],
)
app.include_router(
    fastapi_users.get_register_router(UserRead, UserCreate),
    prefix="/auth",
    tags=["auth"],
)
app.include_router(
    fastapi_users.get_users_router(UserRead, UserUpdate),
    prefix="/users",
    tags=["users"],
)

# Páginas HTML
@app.get("/login.html")
def serve_login():
    return FileResponse("static/login.html")

@app.get("/")
def serve_dashboard():
    return FileResponse("static/index.html")

@app.get("/add.html")
def serve_form():
    return FileResponse("static/add.html")


# ==============================
# API: eficiencias
# ==============================
router = APIRouter(prefix="/api/eficiencias", tags=["eficiencias"])

# Plantilla Excel
@router.get("/template", dependencies=[Depends(get_current_active_user)])
def download_template():
    """
    Genera y envía una plantilla Excel con los encabezados que espera el backend.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    headers = [
        "fecha", "linea", "proceso", "tipo_proceso", "turno", "semana",
        "nombre_asociado", "supervisor", "eficiencia_asociado", "piezas", "tiempo_muerto"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Eficiencias"

    # Encabezados en negritas
    bold = Font(bold=True)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(h) + 2)

    # Fila de ejemplo
    ws.append([
        "2025-07-15", "L28", "Epoxy", "SW", "1er", 29,
        "Juan Pérez", "Ana López", 85.5, 7200, 70
    ])

    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = "plantilla_eficiencias.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )

# Export CSV con filtros
@router.get("/export.csv", dependencies=[Depends(get_current_active_user)])
def export_csv(
    start:   Optional[date] = None,
    end:     Optional[date]  = None,
    linea:   Optional[str]   = None,
    proceso: Optional[str]   = None,
    limit:   int             = 100000,
    db:      Session         = Depends(get_db),
):
    q = db.query(models.Eficiencia)
    if start:   q = q.filter(models.Eficiencia.fecha >= start)
    if end:     q = q.filter(models.Eficiencia.fecha <= end)
    if linea:   q = q.filter(models.Eficiencia.linea == linea)
    if proceso: q = q.filter(models.Eficiencia.proceso == proceso)
    q = q.order_by(models.Eficiencia.fecha.desc()).limit(limit)

    rows: list[models.Eficiencia] = q.all()

    headers = [
        "id", "fecha", "linea", "proceso", "tipo_proceso", "turno", "semana",
        "nombre_asociado", "supervisor", "eficiencia_asociado", "piezas", "tiempo_muerto"
    ]

    sio = StringIO()
    writer = csv.writer(sio)
    writer.writerow(headers)

    for r in rows:
        writer.writerow([
            r.id,
            getattr(r, "fecha", ""),
            getattr(r, "linea", ""),
            getattr(r, "proceso", ""),
            getattr(r, "tipo_proceso", ""),
            getattr(r, "turno", ""),
            getattr(r, "semana", ""),
            getattr(r, "nombre_asociado", ""),
            getattr(r, "supervisor", ""),
            getattr(r, "eficiencia_asociado", ""),
            getattr(r, "piezas", ""),
            getattr(r, "tiempo_muerto", ""),
        ])

    sio.seek(0)
    fname = "eficiencias_filtradas.csv"
    return StreamingResponse(
        iter([sio.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )

# CRUD básico
@router.post(
    "",
    response_model=schemas.Eficiencia,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(get_current_active_admin)],
)
def crear_eficiencia(
    payload: schemas.EficienciaCreate,
    db: Session = Depends(get_db),
):
    reg = models.Eficiencia(**payload.dict())
    db.add(reg)
    db.commit()
    db.refresh(reg)
    return reg

@router.delete(
    "/{ef_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(get_current_active_admin)],
)
def borrar_eficiencia(ef_id: int, db: Session = Depends(get_db)):
    reg = db.query(models.Eficiencia).filter(models.Eficiencia.id == ef_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    db.delete(reg)
    db.commit()

@router.get(
    "",
    response_model=list[schemas.Eficiencia],
    dependencies=[Depends(get_current_active_user)],
)
def leer_eficiencias(
    start:   Optional[date] = None,
    end:     Optional[date]  = None,
    linea:   Optional[str]   = None,
    proceso: Optional[str]   = None,
    skip:    int             = 0,
    limit:   int             = 100,
    db:      Session         = Depends(get_db),
):
    try:
        q = db.query(models.Eficiencia)
        if start:   q = q.filter(models.Eficiencia.fecha >= start)
        if end:     q = q.filter(models.Eficiencia.fecha <= end)
        if linea:   q = q.filter(models.Eficiencia.linea == linea)
        if proceso: q = q.filter(models.Eficiencia.proceso == proceso)
        recientes = (
            q.order_by(models.Eficiencia.id.desc())
             .offset(skip)
             .limit(limit)
             .all()
        )
        return list(reversed(recientes))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Catálogos para filtros
@router.get("/lines", dependencies=[Depends(get_current_active_user)])
def get_lines(db: Session = Depends(get_db)):
    rows = (
        db.query(models.Eficiencia.linea)
          .distinct()
          .order_by(models.Eficiencia.linea)
          .all()
    )
    return [r[0] for r in rows]

@router.get("/processes", dependencies=[Depends(get_current_active_user)])
def get_processes(db: Session = Depends(get_db)):
    rows = (
        db.query(models.Eficiencia.proceso)
          .distinct()
          .order_by(models.Eficiencia.proceso)
          .all()
    )
    return [r[0] for r in rows]

# Métricas
@router.get("/weekly", dependencies=[Depends(get_current_active_user)])
def eficiencia_semanal_por_semana(
    start:   Optional[date] = None,
    end:     Optional[date]  = None,
    linea:   Optional[str]   = None,
    proceso: Optional[str]   = None,
    db:      Session         = Depends(get_db),
):
    q = db.query(
        models.Eficiencia.semana.label("semana"),
        models.Eficiencia.proceso.label("proceso"),
        func.avg(models.Eficiencia.eficiencia_asociado).label("promedio_asociado")
    )
    if start:   q = q.filter(models.Eficiencia.fecha >= start)
    if end:     q = q.filter(models.Eficiencia.fecha <= end)
    if linea:   q = q.filter(models.Eficiencia.linea == linea)
    if proceso: q = q.filter(models.Eficiencia.proceso == proceso)
    q = (
        q.group_by(models.Eficiencia.semana, models.Eficiencia.proceso)
         .order_by(models.Eficiencia.semana)
    )
    return [
        {"semana": int(sem), "proceso": proc, "promedio_asociado": float(round(avg, 2))}
        for sem, proc, avg in q.all()
    ]

@router.get("/daily/process", dependencies=[Depends(get_current_active_user)])
def eficiencia_diaria_proceso(
    start:   Optional[date] = None,
    end:     Optional[date]  = None,
    linea:   Optional[str]   = None,
    proceso: Optional[str]   = None,
    db:      Session         = Depends(get_db),
):
    q = db.query(
        models.Eficiencia.fecha.label("fecha"),
        models.Eficiencia.proceso.label("proceso"),
        func.avg(models.Eficiencia.eficiencia_asociado).label("promedio_asociado")
    )
    if start:   q = q.filter(models.Eficiencia.fecha >= start)
    if end:     q = q.filter(models.Eficiencia.fecha <= end)
    if linea:   q = q.filter(models.Eficiencia.linea == linea)
    if proceso: q = q.filter(models.Eficiencia.proceso == proceso)
    q = (
        q.group_by(models.Eficiencia.fecha, models.Eficiencia.proceso)
         .order_by(models.Eficiencia.fecha)
    )
    return [
        {"fecha": f.isoformat(), "proceso": proc, "promedio_asociado": float(avg)}
        for f, proc, avg in q.all()
    ]

@router.get("/daily/line", dependencies=[Depends(get_current_active_user)])
def eficiencia_diaria_linea(
    start:   Optional[date] = None,
    end:     Optional[date]  = None,
    linea:   Optional[str]   = None,
    proceso: Optional[str]   = None,
    db:      Session         = Depends(get_db),
):
    q = db.query(
        models.Eficiencia.fecha.label("fecha"),
        models.Eficiencia.linea.label("linea"),
        func.avg(models.Eficiencia.eficiencia_asociado).label("promedio_asociado")
    )
    if start:   q = q.filter(models.Eficiencia.fecha >= start)
    if end:     q = q.filter(models.Eficiencia.fecha <= end)
    if linea:   q = q.filter(models.Eficiencia.linea == linea)
    if proceso: q = q.filter(models.Eficiencia.proceso == proceso)
    q = (
        q.group_by(models.Eficiencia.fecha, models.Eficiencia.linea)
         .order_by(models.Eficiencia.fecha)
    )
    return [
        {"fecha": f.isoformat(), "linea": l, "promedio_asociado": float(avg)}
        for f, l, avg in q.all()
    ]

@router.get("/top-associates", dependencies=[Depends(get_current_active_user)])
def worst_associados(
    start:   Optional[date] = None,
    end:     Optional[date]  = None,
    linea:   Optional[str]   = None,
    proceso: Optional[str]   = None,
    limit:   int             = 5,
    db:      Session         = Depends(get_db),
):
    q = db.query(models.Eficiencia)
    if start:   q = q.filter(models.Eficiencia.fecha >= start)
    if end:     q = q.filter(models.Eficiencia.fecha <= end)
    if linea:   q = q.filter(models.Eficiencia.linea == linea)
    if proceso: q = q.filter(models.Eficiencia.proceso == proceso)
    q2 = (
        q.with_entities(
            models.Eficiencia.nombre_asociado.label("nombre"),
            func.avg(models.Eficiencia.eficiencia_asociado).label("promedio_asociado")
        )
        .group_by(models.Eficiencia.nombre_asociado)
        .order_by(func.avg(models.Eficiencia.eficiencia_asociado).asc())
        .limit(limit)
    )
    return [
        {"nombre": nombre, "promedio_asociado": float(round(prom, 2))}
        for nombre, prom in q2.all()
    ]

@router.get("/shift", dependencies=[Depends(get_current_active_user)])
def eficiencia_por_turno(
    start:   Optional[date] = None,
    end:     Optional[date]  = None,
    linea:   Optional[str]   = None,
    proceso: Optional[str]   = None,
    db:      Session         = Depends(get_db),
):
    q = db.query(
        models.Eficiencia.turno.label("turno"),
        func.avg(models.Eficiencia.eficiencia_asociado).label("promedio_asociado")
    )
    if start:   q = q.filter(models.Eficiencia.fecha >= start)
    if end:     q = q.filter(models.Eficiencia.fecha <= end)
    if linea:   q = q.filter(models.Eficiencia.linea == linea)
    if proceso: q = q.filter(models.Eficiencia.proceso == proceso)
    q = (
        q.group_by(models.Eficiencia.turno)
         .order_by(models.Eficiencia.turno)
    )
    return [
        {"turno": t, "promedio_asociado": float(round(p, 2))}
        for t, p in q.all()
    ]

@router.get("/counts", dependencies=[Depends(get_current_active_user)])
def conteo_sw_wd_por_linea(
    start:   Optional[date] = None,
    end:     Optional[date]  = None,
    linea:   Optional[str]   = None,
    proceso: Optional[str]   = None,
    db:      Session         = Depends(get_db),
):
    sub = db.query(
        models.Eficiencia.linea.label("linea"),
        models.Eficiencia.tipo_proceso.label("tipo_proceso"),
        func.sum(models.Eficiencia.piezas).label("total_piezas")
    )
    if start:   sub = sub.filter(models.Eficiencia.fecha >= start)
    if end:     sub = sub.filter(models.Eficiencia.fecha <= end)
    if linea:   sub = sub.filter(models.Eficiencia.linea == linea)
    if proceso: sub = sub.filter(models.Eficiencia.proceso == proceso)
    q = (
        sub.group_by(models.Eficiencia.linea, models.Eficiencia.tipo_proceso)
           .order_by(models.Eficiencia.linea)
    )
    return [
        {"linea": linea, "tipo_proceso": tp, "total_piezas": int(tpz)}
        for linea, tp, tpz in q.all()
    ]

# Downtime semanal por línea (SUM en minutos)
@router.get("/weekly/downtime", dependencies=[Depends(get_current_active_user)])
def downtime_weekly_by_line(
    start:   Optional[date] = None,
    end:     Optional[date]  = None,
    linea:   Optional[str]   = None,
    proceso: Optional[str]   = None,
    db:      Session         = Depends(get_db),
):
    q = db.query(
        models.Eficiencia.semana.label("semana"),
        models.Eficiencia.linea.label("linea"),
        func.sum(models.Eficiencia.tiempo_muerto).label("sum_downtime"),
    )
    if start:   q = q.filter(models.Eficiencia.fecha >= start)
    if end:     q = q.filter(models.Eficiencia.fecha <= end)
    if linea:   q = q.filter(models.Eficiencia.linea == linea)
    if proceso: q = q.filter(models.Eficiencia.proceso == proceso)

    q = (
        q.group_by(models.Eficiencia.semana, models.Eficiencia.linea)
         .order_by(models.Eficiencia.semana)
    )

    return [
        {"semana": int(sem), "linea": line, "sum_downtime": float(dt or 0)}
        for sem, line, dt in q.all()
    ]

# Subida masiva desde Excel (solo admin)
@router.post("/upload", dependencies=[Depends(get_current_active_admin)])
async def upload_eficiencias(
    file: UploadFile = File(...),
    db:   Session    = Depends(get_db),
):
    if not file.filename.lower().endswith(('.xls', '.xlsx')):
        raise HTTPException(400, "Sube un archivo .xls o .xlsx")

    contents = await file.read()
    try:
        sheets = pd.read_excel(io.BytesIO(contents), sheet_name=None)
    except Exception as e:
        raise HTTPException(400, f"Error leyendo el Excel: {e}")

    frames: list[pd.DataFrame] = []
    for sheet_name, df in sheets.items():
        m = re.search(r'(\d{1,2})', sheet_name)
        if not m:
            raise HTTPException(400, f"No encontré número de semana en la hoja '{sheet_name}'")
        semana_val = int(m.group(1))

        df.columns = [c.strip().lower() for c in df.columns]
        required = {
            'nombre_asociado','linea','supervisor',
            'tipo_proceso','proceso','piezas',
            'eficiencia_asociado','turno'
        }
        faltan = required - set(df.columns)
        if faltan:
            raise HTTPException(400, f"Faltan columnas en '{sheet_name}': {', '.join(faltan)}")

        for opt in ('wwid','tiempo_muerto','fecha'):
            if opt not in df.columns:
                df[opt] = None

        df['semana'] = semana_val
        df['piezas'] = pd.to_numeric(df['piezas'], errors='coerce').fillna(0).astype(int)
        df['eficiencia_asociado'] = pd.to_numeric(df['eficiencia_asociado'], errors='coerce').fillna(0.0)
        df['tiempo_muerto'] = pd.to_numeric(df['tiempo_muerto'], errors='coerce')
        if 'fecha' in df.columns:
            df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce').dt.date

        frames.append(df)

    if not frames:
        return {"insertados": 0}

    big_df = pd.concat(frames, ignore_index=True)
    allowed_cols = {c.name for c in models.Eficiencia.__table__.columns}

    registros = [
        models.Eficiencia(**{k: v for k, v in row.items() if k in allowed_cols})
        for row in big_df.to_dict(orient='records')
    ]
    try:
        db.bulk_save_objects(registros)
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(500, "Error insertando en BD. Revisa logs para más detalles.")

    return {"insertados": len(registros)}

# Monta el router principal
app.include_router(router)
